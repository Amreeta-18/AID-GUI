from GMFormInput import GMFormInput
from parseDOM import parseDOM
from CheckRules import CheckRules
from gensim.summarization import keywords
import re
import openpyxl
from openpyxl import load_workbook
import spacy
from collections import Counter
import xlrd
import textParser

# This file runs each of the rules for each URL in the input/output file

def Test(filename):
    wb = openpyxl.load_workbook(filename)
    s = wb.active
    return(s.max_row)

def MainProcess(subgoal, action, filename):

    # filename = "Example_Input.xlsx"      #Name of the file to take input from
    # report = open('report.txt', 'w')  #Output file with the reports
    report = ""

    #Loading the English model for spaCy
    nlp = spacy.load('en_core_web_sm')

    # Initialize an object for the classes imported from other files
    G = GMFormInput()
    # P = parseDOM()
    C = CheckRules()

    res = 0

    # List of DOM words to exclude from keywords
    DOM_words = ['window', 'document', 'header', 'form', 'link', 'field', 'tab', 'button', 'checkbox', 'icon', 'data', 'information']

    # for row in range(2, 3): #The number of rows in the input file to run AID on
    #     keywords_S = []
    #     keywords_A = []
    #     all_keywords = []
    #     # usecase, subgoal, action, url = G.GMFormInput(row, filename) # filename, scenario, subgoal, action, url = G.GMFormInput(row)
    #     # print(row)
    #     # print(url)

    # workbook = openpyxl.load_workbook(filename)
    # sheet = workbook.active

    # Getting Keywords from subgoal and action by extracting nouns
    subgoals = nlp(subgoal)
    actions = nlp(action)
    keywords_S = []
    keywords_A = []
    txt = textParser.textParse(filename)
    link = linkParser.linkParse(filename)

    for token in subgoals:
       if (token.pos_ == 'NOUN' or token.pos_ == 'ADJ') and (str(token) not in DOM_words): #or (token.pos_ == 'NOUN'):
           keywords_S.append(token.text)

    for token in actions:
        if (token.pos_ == 'NOUN' or token.pos_ == 'ADJ') and (str(token) not in DOM_words):
            keywords_A.append(token.text)

    # print(keywords_S,keywords_A)
    # P.get_html(url)
    # filename="current_html.html"
    report = f'URL of webpage evaluated: {filename}\nSubgoal: {subgoal}\nAction: {action}\n'
    
    
   # Rule 1 starts here - tokens present in the page or not
   # print the keywords S and A if violated
    result_1_S = C.checkRule1(keywords_S, txt)
    result_1_A = C.checkRule1(keywords_A, txt)
    if (result_1_S==1 and result_1_A==1):
        # sheet.cell(row+1, 6).value = 0 #"ok"
        report = report + "\n Rule 1 not violated.<br>"
    else:
        # sheet.cell(row+1, 6).value = 1 #"violated"
        report = report + "\n Rule 1 violated: Keywords not found on the webpage.<br>"

    print("Rule 1")

    # #Rule 2 starts here - keywords in 
    # if (row%2==0):
    #     result_2 = C.checkRule2(filename, row)
    #     sheet.cell(row+1, 7).value = result_2
    #     if result_2==1:
    #         report = report + "\n Rule 2 is violated: Keywords from link-label is not present on the current page.\n"
    #     else:
    #         report = report + "\n Rule 2 not violated.\n"
    # else:
    #     sheet.cell(row+1, 7).value = "N/A"
    #     report = report + "\n Rule 2 not applicable.\n"

    # print("Rule 2")


    # #Rule 3 starts here - Link label exists or not
    # Highlight links which donot have label
    # Link labels will be checked
    result_3 = C.checkRule3(filename)
    if result_3==0:
        sheet.cell(row+1, 8).value = 0
        report= report + "\n Rule 3 not violated.\n"
    else:
        sheet.cell(row+1, 8).value = 1
        report = report + "\n Rule 3 violated: Link is not labelled.\n"
    print("Rule 3")
     
    workbook.save(filename)
        
    return report


